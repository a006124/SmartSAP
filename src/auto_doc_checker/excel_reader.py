import polars as pl
import os.path

from sys import exit
from time import sleep
from io import BytesIO
import openpyxl

from auto_doc_checker.auth import get_bq_client

def query_and_fill_excel(excel_file):
    _, file_extension = os.path.splitext(excel_file)
    if not (file_extension == ".xlsx") or not os.path.isfile(excel_file):
        print("File is not Excel or does not exist")
        sleep(5)
        exit(1)

    try:
        query_df = pl.read_excel(source=excel_file, sheet_name="Query")
    except ValueError as e:
        print(repr(e))
        sleep(10)
        exit(1)

    client = get_bq_client(query_df['project_id'][0], query_df['project_location'][0])
    rows = client.query_and_wait(query_df['query'][0])

    query_result_df = pl.from_arrow(rows.to_arrow()) 

    print(query_result_df)

    try:
        buf = BytesIO()
        query_result_df.write_excel(workbook=buf, worksheet="Result")
        buf.seek(0)

        wb = openpyxl.load_workbook(excel_file)
        if "Result" in wb.sheetnames:
            del wb["Result"]
        wb_result = openpyxl.load_workbook(buf)
        ws_new = wb.create_sheet("Result")
        for row in wb_result["Result"].iter_rows(values_only=True):
            ws_new.append(row)
        wb.save(excel_file)
    except PermissionError:
        print(f"Error : cannot write in '{excel_file}', please close the file if opened and try again")
        sleep(10)
        exit(1)
    print("----------------------------------------DONE----------------------------------------")


    